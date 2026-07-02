import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from database.models import User, Purchase, Course


async def export_to_excel(session: AsyncSession) -> io.BytesIO:
    wb = Workbook()

    # ── Sheet 1: Foydalanuvchilar ──
    ws1 = wb.active
    ws1.title = "Foydalanuvchilar"
    headers1 = ["ID", "Telegram ID", "Username", "Ism", "Rol",
                "Balans", "Yechilgan", "Ro'yxatdan o'tgan"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    users = await session.execute(select(User))
    for u in users.scalars():
        ws1.append([
            u.id, u.tg_id, u.username or "", u.full_name,
            u.role, u.balance, u.withdrawn,
            u.registered_at.strftime("%Y-%m-%d %H:%M") if u.registered_at else ""
        ])

    # ── Sheet 2: Xaridlar ──
    ws2 = wb.create_sheet("Xaridlar")
    headers2 = ["ID", "Guruh ID", "Foydalanuvchi ID", "Kurs nomi",
                "Narx", "Asl narx", "Status", "Sana"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="70AD47")
        cell.alignment = Alignment(horizontal="center")

    purchases_q = await session.execute(
        select(Purchase, Course.title)
        .join(Course, Course.id == Purchase.course_id)
        .order_by(Purchase.created_at.desc())
    )
    for p, course_title in purchases_q:
        ws2.append([
            p.id, p.group_id or "", p.user_id, course_title,
            p.amount_paid, p.original_price, p.status,
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else ""
        ])

    # Auto column width
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
